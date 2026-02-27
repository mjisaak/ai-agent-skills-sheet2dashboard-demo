#!/usr/bin/env python3
"""
Sheet-to-Dashboard Sanitizer
Reads an Excel file, cleans and enriches the data, writes sanitized-data.xlsx.

Usage:
    python sanitize.py <input.xlsx> [output.xlsx]

Output:
    sanitized-data.xlsx with two sheets:
      - "data"       : wide format (one row per person)
      - "facts_long" : tidy/long format (one row per person per month)
"""

import sys
import re
import math
from pathlib import Path

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Missing dependencies. Run: pip install pandas openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# City -> Bundesland/Region mapping (DE / AT / CH)
# ---------------------------------------------------------------------------
CITY_BUNDESLAND = {
    # -- Deutschland --
    "berlin": "Berlin",
    "hamburg": "Hamburg",
    "münchen": "Bayern",
    "munich": "Bayern",
    "köln": "Nordrhein-Westfalen",
    "cologne": "Nordrhein-Westfalen",
    "frankfurt": "Hessen",
    "frankfurt am main": "Hessen",
    "stuttgart": "Baden-Württemberg",
    "düsseldorf": "Nordrhein-Westfalen",
    "dortmund": "Nordrhein-Westfalen",
    "essen": "Nordrhein-Westfalen",
    "leipzig": "Sachsen",
    "bremen": "Bremen",
    "dresden": "Sachsen",
    "hannover": "Niedersachsen",
    "nürnberg": "Bayern",
    "duisburg": "Nordrhein-Westfalen",
    "bochum": "Nordrhein-Westfalen",
    "wuppertal": "Nordrhein-Westfalen",
    "bielefeld": "Nordrhein-Westfalen",
    "bonn": "Nordrhein-Westfalen",
    "münster": "Nordrhein-Westfalen",
    "mannheim": "Baden-Württemberg",
    "karlsruhe": "Baden-Württemberg",
    "augsburg": "Bayern",
    "wiesbaden": "Hessen",
    "gelsenkirchen": "Nordrhein-Westfalen",
    "mönchengladbach": "Nordrhein-Westfalen",
    "aachen": "Nordrhein-Westfalen",
    "braunschweig": "Niedersachsen",
    "chemnitz": "Sachsen",
    "kiel": "Schleswig-Holstein",
    "halle": "Sachsen-Anhalt",
    "magdeburg": "Sachsen-Anhalt",
    "freiburg": "Baden-Württemberg",
    "freiburg im breisgau": "Baden-Württemberg",
    "krefeld": "Nordrhein-Westfalen",
    "oberhausen": "Nordrhein-Westfalen",
    "lübeck": "Schleswig-Holstein",
    "erfurt": "Thüringen",
    "rostock": "Mecklenburg-Vorpommern",
    "mainz": "Rheinland-Pfalz",
    "kassel": "Hessen",
    "hagen": "Nordrhein-Westfalen",
    "potsdam": "Brandenburg",
    "saarbrücken": "Saarland",
    "hamm": "Nordrhein-Westfalen",
    "mülheim": "Nordrhein-Westfalen",
    "ludwigshafen": "Rheinland-Pfalz",
    "oldenburg": "Niedersachsen",
    "osnabrück": "Niedersachsen",
    "leverkusen": "Nordrhein-Westfalen",
    "darmstadt": "Hessen",
    "heidelberg": "Baden-Württemberg",
    "solingen": "Nordrhein-Westfalen",
    "herne": "Nordrhein-Westfalen",
    "neuss": "Nordrhein-Westfalen",
    "regensburg": "Bayern",
    "ingolstadt": "Bayern",
    "würzburg": "Bayern",
    "wolfsburg": "Niedersachsen",
    "ulm": "Baden-Württemberg",
    "göttingen": "Niedersachsen",
    "pforzheim": "Baden-Württemberg",
    "offenbach": "Hessen",
    "bottrop": "Nordrhein-Westfalen",
    "bremerhaven": "Bremen",
    "recklinghausen": "Nordrhein-Westfalen",
    "remscheid": "Nordrhein-Westfalen",
    "fürth": "Bayern",
    "trier": "Rheinland-Pfalz",
    "koblenz": "Rheinland-Pfalz",
    "erlangen": "Bayern",
    "moers": "Nordrhein-Westfalen",
    "siegen": "Nordrhein-Westfalen",
    "hildesheim": "Niedersachsen",
    "jena": "Thüringen",
    # -- Österreich --
    "wien": "Wien",
    "vienna": "Wien",
    "graz": "Steiermark",
    "linz": "Oberösterreich",
    "salzburg": "Salzburg",
    "innsbruck": "Tirol",
    "klagenfurt": "Kärnten",
    "villach": "Kärnten",
    "wels": "Oberösterreich",
    "st. pölten": "Niederösterreich",
    "dornbirn": "Vorarlberg",
    "wiener neustadt": "Niederösterreich",
    "steyr": "Oberösterreich",
    "feldkirch": "Vorarlberg",
    "bregenz": "Vorarlberg",
    "leonding": "Oberösterreich",
    "leoben": "Steiermark",
    "amstetten": "Niederösterreich",
    # -- Schweiz --
    "zürich": "Zürich",
    "zurich": "Zürich",
    "genf": "Genf",
    "genève": "Genf",
    "basel": "Basel-Stadt",
    "bern": "Bern",
    "lausanne": "Waadt",
    "winterthur": "Zürich",
    "luzern": "Luzern",
    "st. gallen": "St. Gallen",
    "biel": "Bern",
    "thun": "Bern",
    "köniz": "Bern",
    "la chaux-de-fonds": "Neuenburg",
    "schaffhausen": "Schaffhausen",
    "fribourg": "Freiburg",
    "freiburg": "Freiburg",
    "chur": "Graubünden",
    "vernier": "Genf",
    "uster": "Zürich",
    "sion": "Wallis",
    "neuchâtel": "Neuenburg",
    "lugano": "Tessin",
    "zug": "Zug",
    "aarau": "Aargau",
    "emmen": "Luzern",
}


def lookup_bundesland(city: str) -> str:
    """Map city name to Bundesland/Kanton/Land."""
    key = str(city).strip().lower()
    return CITY_BUNDESLAND.get(key, "Unbekannt")


def detect_umsatz_cols(df: pd.DataFrame):
    """Return sorted list of Umsatz_YYYY-MM column names."""
    pattern = re.compile(r"^Umsatz_\d{4}-\d{2}$", re.IGNORECASE)
    cols = [c for c in df.columns if pattern.match(str(c))]
    cols.sort()
    return cols


def normalize_name(df: pd.DataFrame, warnings: list) -> pd.DataFrame:
    """Split Name -> Vorname / Nachname if needed, trim existing cols."""
    cols = [c.strip() for c in df.columns]
    df.columns = [c.strip() for c in df.columns]

    has_name = "Name" in df.columns
    has_vorname = "Vorname" in df.columns
    has_nachname = "Nachname" in df.columns

    if has_vorname and has_nachname:
        df["Vorname"] = df["Vorname"].astype(str).str.strip()
        df["Nachname"] = df["Nachname"].astype(str).str.strip()
    elif has_name:
        def split_name(val):
            parts = str(val).strip().split()
            parts = [p for p in parts if p]
            if len(parts) == 0:
                return ("", "")
            elif len(parts) == 1:
                return (parts[0], "")
            else:
                # First token = Vorname, rest = Nachname (handles double surnames)
                return (parts[0], " ".join(parts[1:]))

        split = df["Name"].apply(split_name)
        df.insert(0, "Vorname", split.apply(lambda x: x[0]))
        df.insert(1, "Nachname", split.apply(lambda x: x[1]))
        df.drop(columns=["Name"], inplace=True)
    else:
        warnings.append("WARN: Keine Name/Vorname/Nachname-Spalte gefunden. Spalten werden als leer erzeugt.")
        df.insert(0, "Vorname", "")
        df.insert(1, "Nachname", "")

    return df


def schema_check(df: pd.DataFrame) -> list:
    """Return list of error/warning strings for missing expected columns."""
    required = {"Stadt", "Beruf", "Abteilung", "Teilzeit", "Alter"}
    issues = []
    for col in required:
        if col not in df.columns:
            issues.append(f"ERROR: Pflicht-Spalte '{col}' fehlt. Bitte Spalte hinzufügen oder umbenennen.")
    return issues


def sanitize(input_path: str, output_path: str):
    warnings = []
    errors = []

    print(f"Lese: {input_path}")
    df = pd.read_excel(input_path, sheet_name=0)
    df.columns = [str(c).strip() for c in df.columns]
    original_count = len(df)

    # 1. Schema check
    issues = schema_check(df)
    for issue in issues:
        if issue.startswith("ERROR"):
            errors.append(issue)
            print(issue)
    if errors:
        print(f"\n{len(errors)} Fehler gefunden. Bitte Daten korrigieren.")
        sys.exit(1)

    # 2. Name normalisieren
    df = normalize_name(df, warnings)

    # 3. Datentypen normalisieren
    #    Alter -> int
    df["Alter"] = pd.to_numeric(df["Alter"], errors="coerce").fillna(0).astype(int)

    #    Teilzeit -> Ja/Nein
    def norm_teilzeit(val):
        v = str(val).strip().lower()
        if v in ("ja", "yes", "1", "true", "j", "y"):
            return "Ja"
        elif v in ("nein", "no", "0", "false", "n"):
            return "Nein"
        else:
            warnings.append(f"WARN: Unbekannter Teilzeit-Wert '{val}' -> 'Nein'")
            return "Nein"

    df["Teilzeit"] = df["Teilzeit"].apply(norm_teilzeit)

    #    Stadt / Bundesland
    df["Stadt"] = df["Stadt"].astype(str).str.strip()

    #    Umsatz_* columns
    umsatz_cols = detect_umsatz_cols(df)
    if not umsatz_cols:
        warnings.append("WARN: Keine Umsatz_YYYY-MM-Spalten gefunden.")

    for col in umsatz_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        df[col] = df[col].clip(lower=0)

    # 4. Bundesland ableiten
    df["Bundesland"] = df["Stadt"].apply(lookup_bundesland)
    unknown_cities = df[df["Bundesland"] == "Unbekannt"]["Stadt"].unique().tolist()
    if unknown_cities:
        warnings.append(f"WARN: Unbekannte Städte ({len(unknown_cities)}): {', '.join(unknown_cities[:10])}")

    # 5. Umsatz-Features berechnen
    if umsatz_cols:
        df["Umsatz_Gesamt"] = df[umsatz_cols].sum(axis=1)
        df["Umsatz_OE_Monat"] = (df["Umsatz_Gesamt"] / len(umsatz_cols)).round(2)
    else:
        df["Umsatz_Gesamt"] = 0
        df["Umsatz_OE_Monat"] = 0

    # 6. Spaltenreihenfolge + Sortierung
    base_cols = ["Vorname", "Nachname", "Stadt", "Bundesland", "Abteilung", "Beruf", "Teilzeit", "Alter"]
    tail_cols = ["Umsatz_Gesamt", "Umsatz_OE_Monat"]
    extra_cols = [c for c in df.columns if c not in base_cols + umsatz_cols + tail_cols]
    # drop original Name if still present
    extra_cols = [c for c in extra_cols if c != "Name"]

    ordered_cols = base_cols + umsatz_cols + tail_cols
    df = df[ordered_cols]
    df = df.sort_values(["Abteilung", "Beruf", "Nachname"]).reset_index(drop=True)

    # 7. Long format (tidy)
    id_cols = ["Vorname", "Nachname", "Stadt", "Bundesland", "Abteilung", "Beruf", "Teilzeit", "Alter"]
    if umsatz_cols:
        df_long = df[id_cols + umsatz_cols].melt(
            id_vars=id_cols,
            value_vars=umsatz_cols,
            var_name="Datum",
            value_name="Umsatz"
        )
        df_long["Datum"] = df_long["Datum"].str.replace("Umsatz_", "", regex=False)
        df_long = df_long.sort_values(["Abteilung", "Nachname", "Datum"]).reset_index(drop=True)
    else:
        df_long = pd.DataFrame()

    # 8. Summary
    date_range = f"{umsatz_cols[0].replace('Umsatz_','')} bis {umsatz_cols[-1].replace('Umsatz_','')}" if umsatz_cols else "keine"
    print(f"\n--- Zusammenfassung ---")
    print(f"Datensaetze:       {original_count}")
    print(f"Umsatzmonate:      {len(umsatz_cols)} ({date_range})")
    print(f"Warnungen:         {len(warnings)}")
    for w in warnings:
        print(f"  {w}")
    print(f"Default-Filter:    Alle Abteilungen, letztes volles Jahr")

    # 9. Write Excel
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="data", index=False)
        if not df_long.empty:
            df_long.to_excel(writer, sheet_name="facts_long", index=False)

        # Style header rows
        wb = writer.book
        for sheet_name in writer.sheets:
            ws = wb[sheet_name]
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2D4A6B")
                cell.alignment = Alignment(horizontal="center")
            ws.freeze_panes = "A2"

    print(f"\nGespeichert: {output_path}")
    return df, df_long, umsatz_cols, warnings


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python sanitize.py <input.xlsx> [output.xlsx]")
        sys.exit(1)
    inp = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else "sanitized-data.xlsx"
    sanitize(inp, out)
